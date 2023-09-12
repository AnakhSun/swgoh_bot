import json
import requests
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
# Считываем данные из JSON файла
guild_url = 'https://swgoh.gg/g/Z0kME2OMScC4ipNLpCpeSw/'


def add_user():
    pass

wb = Workbook()

# Удаляем начальную страницу (по умолчанию sheet)
default_sheet = wb['Sheet']


# Отправляем GET-запрос и создаем объект BeautifulSoup
guild_response = requests.get(guild_url)
guild_soup = BeautifulSoup(guild_response.text, 'html.parser')

soup = guild_soup.select('body > div.container.p-t-md > div.content-container > div.content-container-primary.character-list > ul > li.media.list-group-item.p-0.b-t-0 > div > table > tbody')

table_rows = guild_soup.find_all('tr')

data_list = {}

for row in table_rows:
    data = row.find_all('td')
    if len(data) > 1:
        nickname = data[0].find('strong').text
        gp = int(data[1].text)
        rank = data[2].text.strip()
        wins = data[3].text.strip()
        losses = data[4].text.strip()
        member_status = data[5].text.strip()

        # Создайте словарь для текущей строки
        row_data = {
            #"Nickname": username,
            "GP": gp,
            "Rank": rank,
            "Arena": wins,
            "Fleet": losses,
            "Role": member_status
        }

        # Добавьте словарь в список
        data_list[nickname] = row_data

player_roles = {}
with open('C:/Users/user/Desktop/swgoh_bot/data_swgoh_332.json', 'r', encoding='utf-8') as json_file:
    data = json.load(json_file)

merged_data = {}

# Пройдитесь по данным из второго файла и объедините их с данными из первого файла по имени
for key2, value2 in data.items():
    player_name = value2["player_name"]
    if player_name in data_list:
        data_list[player_name].update(value2)
        merged_data[key2] = data_list[player_name]



if merged_data != {}:
    data = merged_data
    with open('C:/Users/user/Desktop/swgoh_bot/data_swgoh_332.json', 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)



# Проходим по данным из JSON файла и записываем их в файл Excel
for player_id, player_data in data.items():
    # Получаем имя игрока с сайта
    player_url = f'https://swgoh.gg/p/{player_id}/'
    response = requests.get(player_url)
    #role = player_roles.get(player_name, '')
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        player_name = soup.find('h5').text.strip()
        galactic_power_element = soup.select_one('html > body > div:nth-of-type(3) > div:nth-of-type(1) > div:nth-of-type(2) > ul > li:nth-of-type(8) > div > div:nth-of-type(1) > span:nth-of-type(2) > strong')
        galactic_power = galactic_power_element.text
        lvl = soup.select_one('body > div.container.p-t-md > div.content-container > div.content-container-aside > div.panel.panel-default.panel-profile.m-b-sm > div.panel-body > ul > li:nth-child(3) > h5').text.strip()
        # Удалите запятые и конвертируйте в число
        galactic_power = int(galactic_power.replace(',', ''))
    else:
        player_name = 'Unknown'
        galactic_power = 0  # По умолчанию

    # Добавляем имя игрока и галактическую мощь в JSON данные
    player_data['player_name'] = player_name
    player_data['galactic_power'] = galactic_power

    # Создаем новый лист с именем недели
    progress_in_week = player_data.get('progress_in_week', {})
    for week, week_data in progress_in_week.items():
        avg_energy = week_data.get('avg_energy', '')
        activ_gild_war = week_data.get('activ_gild_war', '')
        activ_battles = week_data.get('activ_battles', '')
        role = player_data.get('Role', '')
        lvl = player_data.get('lvl', '')

        # Создаем новый лист с именем недели
        ws = wb.create_sheet(title=week)

        # Записываем заголовки столбцов
        headers = ["Player Name", "Galactic Power", "Player ID", "Level", "Role", "Energy", "GW", "TW", "Plan"]
        ws.append(headers)
        # Записываем данные в файл Excel
        row = [player_name, galactic_power, player_id, lvl, role, avg_energy, activ_gild_war, activ_battles, player_data.get('plan', '')]
        print(data_list.get('Role', ''))
        ws.append(row)

# Обновляем исходный JSON файл с добавленными данными
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get the column letter (e.g., 'A', 'B', 'C', ...)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add a little extra padding
            ws.column_dimensions[column_letter].width = adjusted_width

# Сохраняем файл Excel
wb.remove(default_sheet)
wb.save('output.xlsx')

print("Данные успешно записаны в файл Excel и обновлены в исходном JSON файле.")